from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from robots.models import Robot
from orders.views import find_order
from django.core import serializers

import json
from openpyxl import Workbook
from datetime import timedelta, datetime
from django.utils import timezone
from django.db.models import Count, Q


@csrf_exempt
def add_robot(request):
    if request.method == 'POST':
        info = json.loads(request.body.decode("utf-8"))
        info['created'] = datetime.strptime(info['created'], '%Y-%m-%d %H:%M:%S') #.strftime(info['created'], '%Y-%m-%d %H:%M:%S'))
        robot = Robot(**info)

        if robot.save():
            find_order(info.serial)
            return HttpResponse(201)
        return HttpResponse(304)
    return HttpResponse(400)


def download_excel(request):
    if request.method == 'GET':
        robots = Robot.objects.all()
        serialized_data = serializers.serialize('json', robots)
        python_objects = json.loads(serialized_data)
        print(python_objects)
        '''
        formatted_data = []

        for instance in robots:

            formatted_data.append(formatted_instance)

        formatted_json = serializers.serialize('json', formatted_data)
        print(robots)
        '''

        wb = Workbook()
        header = [f.name.capitalize() for f in Robot._meta.get_fields()]
        for sheet_name in set(Robot.objects.values_list('model')):
            ws = wb.create_sheet(sheet_name[0])
            for index, item in enumerate(header, 1):
                cell = ws.cell(row=1, column=index)
                cell.value = item

            today = datetime.now(tz=timezone.utc).strftime("%d-%m-%Y %H:%M:%S")
            start = datetime.now(tz=timezone.utc) - timedelta(days=datetime.now(tz=timezone.utc).weekday())
            end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)

             #.annotate(version=Count('version'))# #

            for robot in Robot.objects.all():
                robot.created = Robot.objects.filter(created__gte=start, created__lte=end).count()
                formatted_instance = [
                    robot.model,
                    robot.version,
                    robot.created,  # Format the datetime
                ]
                ws.append(formatted_instance)

        wb.save('output.xlsx')

        return HttpResponse(robot)

        # for row in ws.iter_rows(min_row=1, max_col=len(header)):
           #     pass



        # ws = wb.active


